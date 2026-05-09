#!/usr/bin/env python3
"""
Bordkasse v8 → v9 Migration

Drei Änderungen:
  1. Gutschrift wandert in einen eigenen Tab (raus aus Eingabe).
  2. Eingabe-Tab kompakter (Art-Toggle entfernt → ~26% weniger Scrollen mobil).
  3. Apps Script bekommt Datum-Autofill (separate Datei _v9.js).

Aufruf:
  python3 scripts/migrate_v8_to_v9.py

Erzeugt:
  assets/sheets-current/Bordkasse_IJsselmeer2026_v9.xlsx
  assets/sheets-current/Bordkasse_AppScript_v9.js

Original v8-Dateien bleiben unverändert.
"""

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

ROOT  = Path(__file__).resolve().parent.parent
SHEET_DIR = ROOT / "assets" / "sheets-current"
SRC_XLSX  = SHEET_DIR / "Bordkasse_IJsselmeer2026.xlsx"
DST_XLSX  = SHEET_DIR / "Bordkasse_IJsselmeer2026_v9.xlsx"
DST_JS    = SHEET_DIR / "Bordkasse_AppScript_v9.js"

# Welche Original-Zeilen werden aus Eingabe gelöscht?
#   4–7  : "1. Was ist es?" + Art-Label + Art-Dropdown + Spacer
#   48–54: kompletter alter Gutschrift-Block (Header + Von + An + Hint)
DELETIONS = list(range(4, 8)) + list(range(48, 55))


def shift_row(orig_row: int) -> int:
    """Liefert die neue Zeilennummer einer Zelle nach allen Löschungen."""
    return orig_row - sum(1 for d in DELETIONS if d < orig_row)


def clone_style(src_cell, dst_cell):
    """Kopiert Font/Fill/Border/Alignment/Format zwischen Zellen."""
    if src_cell.has_style:
        dst_cell.font          = copy(src_cell.font)
        dst_cell.fill          = copy(src_cell.fill)
        dst_cell.border        = copy(src_cell.border)
        dst_cell.alignment     = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection    = copy(src_cell.protection)


def add_dv(ws, formula1, target, dv_type="list", allow_blank=True):
    """Hängt eine neue Data-Validation an die Zielzelle/-bereich an."""
    dv = DataValidation(type=dv_type, formula1=formula1, allow_blank=allow_blank)
    dv.add(target)
    ws.add_data_validation(dv)


# ─────────────────────────────────────────────────────────────────────────
# 1) Eingabe-Tab umbauen
# ─────────────────────────────────────────────────────────────────────────
def restructure_eingabe(wb):
    ws = wb["Eingabe"]

    # Original-Merges sichern, dann alle entfernen
    orig_merges = [(mr.min_row, mr.max_row, mr.min_col, mr.max_col)
                   for mr in list(ws.merged_cells.ranges)]
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Alle bestehenden Data-Validations verwerfen (waren teils stale)
    ws.data_validations = DataValidationList()

    # Zeilen löschen — von unten nach oben, damit Indizes stabil bleiben
    ws.delete_rows(48, 7)   # alter Gutschrift-Block (rows 48–54)
    ws.delete_rows(4, 4)    # alter Art-Block (rows 4–7)

    # Subtitle aktualisieren
    ws["A2"].value = "Ausgabe erfassen"

    # Section-Header umnummerieren (nach der Löschung):
    #   "  1.  Was ist es?"  ist weg
    #   "  2.  Details"      → "  1.  Details"     (jetzt A4)
    #   "  3.  Aufteilung"   → "  2.  Aufteilung"  (jetzt A23)
    if ws["A4"].value and "Details" in str(ws["A4"].value):
        ws["A4"].value = "  1.  Details"
    if ws["A23"].value and "Aufteilung" in str(ws["A23"].value):
        ws["A23"].value = "  2.  Aufteilung"

    # Merges neu anlegen — verschoben um die gelöschten Zeilen
    for (mn, mx, cn, cx) in orig_merges:
        # Merges, die innerhalb gelöschter Zeilen lagen, fallen weg
        if any(mn <= d <= mx for d in DELETIONS):
            continue
        new_min = shift_row(mn)
        new_max = shift_row(mx)
        ws.merge_cells(start_row=new_min, end_row=new_max,
                       start_column=cn,   end_column=cx)

    # Data-Validations frisch setzen (an neuen Koordinaten)
    add_dv(ws, "Kategorien!$B$4:$B$23",                                  "B12")  # Kategorie
    add_dv(ws, "Besatzung!$B$11:$B$22",                                  "B15")  # Bezahlt von
    add_dv(ws, '"Gleichmäßig,An Bord,Zeitanteilig,Individuell"',         "B25")  # Aufteilung
    add_dv(ws, '"x"',                                                    "D31:D42")  # Anwesenheit


# ─────────────────────────────────────────────────────────────────────────
# 2) Gutschrift-Tab neu erstellen
# ─────────────────────────────────────────────────────────────────────────
def build_gutschrift(wb):
    if "Gutschrift" in wb.sheetnames:
        del wb["Gutschrift"]

    eingabe = wb["Eingabe"]
    ws = wb.create_sheet("Gutschrift", index=1)  # direkt nach Eingabe

    # Spaltenbreiten von Eingabe übernehmen
    for letter, dim in eingabe.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[letter].width = dim.width

    # Style-Quellen (nach Eingabe-Umbau gültige Koordinaten)
    style_title    = eingabe["A1"]   # ⚓ Bordkasse
    style_subtitle = eingabe["A2"]   # "Ausgabe erfassen"
    style_section  = eingabe["A4"]   # "1. Details"
    style_label    = eingabe["B5"]   # "Datum"
    style_field    = eingabe["B6"]   # Datum value
    style_hint     = eingabe["A41"] if eingabe["A41"].value else eingabe["B7"]
    style_save     = eingabe["A40"]  # SPEICHERN
    style_save_hint= eingabe["A41"]  # Hint unter Speichern

    # Layout der Gutschrift
    rows = [
        # (zeile, höhe, zelle, wert, style_quelle, merge_bis_spalte)
        (1,  55.5,  "A1", "⚓  Bordkasse",                           style_title,    "C"),
        (2,  24.0,  "A2", "Gutschrift erfassen",                     style_subtitle, "C"),
        (3,  12.0,  None, None,                                      None,           None),
        (4,  33.75, "A4", "  1.  Details",                           style_section,  "C"),
        (5,  21.75, "B5", "Datum",                                   style_label,    None),
        (6,  43.5,  "B6", None,                                      style_field,    None),
        (7,  7.5,   None, None,                                      None,           None),
        (8,  21.75, "B8", "Beschreibung  (optional)",                style_label,    None),
        (9,  43.5,  "B9", None,                                      style_field,    None),
        (10, 7.5,   None, None,                                      None,           None),
        (11, 21.75, "B11","Betrag  (€)",                             style_label,    None),
        (12, 49.5,  "B12",None,                                      style_field,    None),
        (13, 13.5,  None, None,                                      None,           None),
        (14, 33.75, "A14","  2.  Wer hat wem gezahlt?",              style_section,  "C"),
        (15, 21.75, "B15","Zahlt  (Von)",                            style_label,    None),
        (16, 43.5,  "B16",None,                                      style_field,    None),
        (17, 7.5,   None, None,                                      None,           None),
        (18, 21.75, "B18","Empfängt  (An)",                          style_label,    None),
        (19, 43.5,  "B19",None,                                      style_field,    None),
        (20, 21.75, "B20","Name eines Crewmitglieds oder 'Alle' "
                          "(Aufteilung an gesamte Crew)",            style_hint,     None),
        (21, 21.75, None, None,                                      None,           None),
        (22, 60.0,  "A22","▶   GUTSCHRIFT SPEICHERN",                style_save,     "E"),
        (23, 21.75, "A23","Tippe hier (Desktop) oder nutze Menü: "
                          "⚓ Bordkasse → 💾 Gutschrift speichern",   style_save_hint,"E"),
    ]

    for (row, height, cell_ref, value, style_src, merge_to) in rows:
        ws.row_dimensions[row].height = height
        if cell_ref is None:
            continue
        ws[cell_ref].value = value
        if style_src is not None:
            clone_style(style_src, ws[cell_ref])
        if merge_to is not None:
            ws.merge_cells(f"A{row}:{merge_to}{row}")

    # Format für Datum + Betrag setzen
    ws["B6"].number_format  = "DD.MM.YYYY"
    ws["B12"].number_format = '#,##0.00 "€"'

    # Data-Validations
    add_dv(ws, "Besatzung!$B$11:$B$22",  "B16")  # Zahlt (Von)
    add_dv(ws, "Besatzung!$K$11:$K$23",  "B19")  # Empfängt (An, inkl. "Alle")


# ─────────────────────────────────────────────────────────────────────────
# 3) Apps Script v9 schreiben
# ─────────────────────────────────────────────────────────────────────────
APPS_SCRIPT_V9 = r"""// ═══════════════════════════════════════════════════════════════════════
// BORDKASSE IJsselmeer 2026 — Apps Script v9
//
// NEU in v9:
// - Gutschrift hat eigenen Tab (eigenständige Maske + gutschriftSpeichern())
// - Eingabe-Tab kompakter — Art-Toggle entfernt, neue Zellkoordinaten
// - Datum-Autofill: beim Leeren / Öffnen wird Datum auf heute gesetzt
//
// INSTALLATION (einmalig):
//   1. Google Sheets öffnen → Erweiterungen → Apps Script
//   2. Vorhandenen Code markieren (Strg+A) und löschen
//   3. Diesen kompletten Code einfügen
//   4. Speichern (Strg+S)
//   5. Tabelle neu laden (F5)
// ═══════════════════════════════════════════════════════════════════════

// ── Eingabe (Ausgaben) ─────────────────────────────────────────────────
const E = {
  SHEET:        "Eingabe",
  DATUM:        "B6",
  BESCHREIBUNG: "B9",
  KATEGORIE:    "B12",
  BEZAHLT_VON:  "B15",
  BETRAG:       "B18",
  ALKOHOL:      "B21",
  AUFTEILUNG:   "B25",
  DABEI_START_ROW: 31,
  DABEI_COL:       4,
  DABEI_COUNT:     12,
  INDIVIDUELL_FIRST_ROW: 28,
  INDIVIDUELL_LAST_ROW:  42,
};

// ── Gutschrift (eigener Tab) ───────────────────────────────────────────
const G = {
  SHEET:        "Gutschrift",
  DATUM:        "B6",
  BESCHREIBUNG: "B9",
  BETRAG:       "B12",
  GUT_VON:      "B16",
  GUT_AN:       "B19",
};

// ── Transaktionen ──────────────────────────────────────────────────────
const TX = {
  SHEET:    "Transaktionen",
  START:    4,
  MAX:      203,
  DATUM:    2,  TYP: 3,  DESC: 4,  KAT: 5,
  PAID:     6,  BETRAG: 7,  ALKOHOL: 8,  SPLIT: 9,
  DABEI_S:  10,
  GUT_VON:  34,
  GUT_AN:   35,
};

const B = {
  SHEET:      "Besatzung",
  NAME_COL:   2,
  NAME_START: 11,
  N:          12,
};

const S = {
  SHEET:      "Schulden",
  DATA_START: 8,
  DATA_MAX:   27,
  TIMESTAMP:  31,
};


// ════════════════════════════════════════════════════════════════════════
// AUTO-TRIGGER: Aufteilung B25 → Individuell-Block toggeln
// ════════════════════════════════════════════════════════════════════════
function onEdit(e) {
  if (!e || !e.range) return;
  const sheet = e.range.getSheet();
  if (sheet.getName() !== E.SHEET) return;

  if (e.range.getA1Notation() === E.AUFTEILUNG) {
    toggleIndividuellBlock_(sheet);
  }
}


function toggleIndividuellBlock_(wsEin) {
  const aufteilung = (wsEin.getRange(E.AUFTEILUNG).getValue() || "").toString().trim();
  const firstRow = E.INDIVIDUELL_FIRST_ROW;
  const numRows  = E.INDIVIDUELL_LAST_ROW - firstRow + 1;

  if (aufteilung === "Individuell") {
    wsEin.showRows(firstRow, numRows);
  } else {
    wsEin.hideRows(firstRow, numRows);
    for (let i = 0; i < E.DABEI_COUNT; i++) {
      wsEin.getRange(E.DABEI_START_ROW + i, E.DABEI_COL).setValue("");
    }
  }
}


// ════════════════════════════════════════════════════════════════════════
// AUSGABE SPEICHERN (Eingabe-Tab)
// ════════════════════════════════════════════════════════════════════════
function transaktionSpeichern() {
  const ss    = SpreadsheetApp.getActiveSpreadsheet();
  const wsEin = ss.getSheetByName(E.SHEET);
  const wsTx  = ss.getSheetByName(TX.SHEET);
  const ui    = SpreadsheetApp.getUi();

  if (!wsEin || !wsTx) {
    ui.alert("❌ Fehler", "Tabellenblätter 'Eingabe' oder 'Transaktionen' nicht gefunden.", ui.ButtonSet.OK);
    return;
  }

  const datum        = wsEin.getRange(E.DATUM).getValue();
  const beschreibung = (wsEin.getRange(E.BESCHREIBUNG).getValue() || "").toString().trim();
  const kategorie    = (wsEin.getRange(E.KATEGORIE).getValue() || "").toString().trim();
  const bezahltVon   = (wsEin.getRange(E.BEZAHLT_VON).getValue() || "").toString().trim();
  const betrag       = wsEin.getRange(E.BETRAG).getValue();
  const alkohol      = wsEin.getRange(E.ALKOHOL).getValue() || 0;
  const aufteilung   = (wsEin.getRange(E.AUFTEILUNG).getValue() || "Gleichmäßig").toString().trim();

  const dabei = [];
  for (let i = 0; i < E.DABEI_COUNT; i++) {
    const row = E.DABEI_START_ROW + i;
    const val = (wsEin.getRange(row, E.DABEI_COL).getValue() || "").toString().trim();
    dabei.push(val);
  }

  if (!beschreibung) return alertMissing_(ui, wsEin, E.BESCHREIBUNG, "Bitte eine Beschreibung eingeben.");
  if (!betrag || Number(betrag) === 0) return alertMissing_(ui, wsEin, E.BETRAG, "Bitte einen Betrag eingeben.");
  if (!bezahltVon) return alertMissing_(ui, wsEin, E.BEZAHLT_VON, "Bitte angeben, wer bezahlt hat.");
  if (alkohol && Number(alkohol) > Number(betrag)) {
    return alertMissing_(ui, wsEin, E.ALKOHOL, "Alkohol-Anteil darf nicht größer als der Gesamtbetrag sein.");
  }

  const nextRow = freieZeileFinden_(wsTx);
  if (nextRow === -1) {
    ui.alert("❌ Liste voll", `Max. ${TX.MAX - TX.START + 1} Einträge erreicht.`, ui.ButtonSet.OK);
    return;
  }

  let datumWert = (datum && datum.toString().trim() !== "") ? datum : new Date();

  wsTx.getRange(nextRow, TX.DATUM).setValue(datumWert).setNumberFormat("DD.MM.YYYY");
  wsTx.getRange(nextRow, TX.TYP).setValue("Ausgabe");
  wsTx.getRange(nextRow, TX.DESC).setValue(beschreibung);
  wsTx.getRange(nextRow, TX.KAT).setValue(kategorie);
  wsTx.getRange(nextRow, TX.PAID).setValue(bezahltVon);
  wsTx.getRange(nextRow, TX.BETRAG).setValue(betrag).setNumberFormat('#,##0.00 "€"');

  if (alkohol && Number(alkohol) > 0) {
    wsTx.getRange(nextRow, TX.ALKOHOL).setValue(alkohol).setNumberFormat('#,##0.00 "€"');
  }

  wsTx.getRange(nextRow, TX.SPLIT).setValue(aufteilung);

  if (aufteilung === "Individuell") {
    for (let p = 0; p < B.N; p++) {
      if (dabei[p] && dabei[p].toLowerCase() === "x") {
        wsTx.getRange(nextRow, TX.DABEI_S + p).setValue("x");
      }
    }
  }

  eingabeMaskeLeeren_(wsEin);
  schuldenBerechnen_(ss);

  const betragText = Number(betrag).toLocaleString("de-DE", {
    minimumFractionDigits: 2, maximumFractionDigits: 2
  }) + " €";

  ui.alert("✅ Gespeichert!",
    `Ausgabe in Zeile ${nextRow} eingetragen.\n\n` +
    `Beschreibung: ${beschreibung}\nBetrag: ${betragText}\nBezahlt von: ${bezahltVon}`,
    ui.ButtonSet.OK);
  wsEin.getRange(E.DATUM).activate();
}


// ════════════════════════════════════════════════════════════════════════
// GUTSCHRIFT SPEICHERN (eigener Tab)
// ════════════════════════════════════════════════════════════════════════
function gutschriftSpeichern() {
  const ss    = SpreadsheetApp.getActiveSpreadsheet();
  const wsGut = ss.getSheetByName(G.SHEET);
  const wsTx  = ss.getSheetByName(TX.SHEET);
  const ui    = SpreadsheetApp.getUi();

  if (!wsGut || !wsTx) {
    ui.alert("❌ Fehler", "Tabellenblätter 'Gutschrift' oder 'Transaktionen' nicht gefunden.", ui.ButtonSet.OK);
    return;
  }

  const datum        = wsGut.getRange(G.DATUM).getValue();
  const beschreibung = (wsGut.getRange(G.BESCHREIBUNG).getValue() || "").toString().trim();
  const betrag       = wsGut.getRange(G.BETRAG).getValue();
  const gutVon       = (wsGut.getRange(G.GUT_VON).getValue() || "").toString().trim();
  const gutAn        = (wsGut.getRange(G.GUT_AN).getValue() || "").toString().trim();

  if (!gutVon || !gutAn) return alertMissing_(ui, wsGut, G.GUT_VON, "Bitte 'Zahlt (Von)' und 'Empfängt (An)' ausfüllen.");
  if (gutVon === gutAn) {
    ui.alert("⚠️ Fehler", "'Von' und 'An' können nicht dieselbe Person sein.", ui.ButtonSet.OK);
    return;
  }
  if (!betrag || Number(betrag) === 0) return alertMissing_(ui, wsGut, G.BETRAG, "Bitte den Betrag der Gutschrift eingeben.");

  const nextRow = freieZeileFinden_(wsTx);
  if (nextRow === -1) {
    ui.alert("❌ Liste voll", `Max. ${TX.MAX - TX.START + 1} Einträge erreicht.`, ui.ButtonSet.OK);
    return;
  }

  let datumWert = (datum && datum.toString().trim() !== "") ? datum : new Date();

  wsTx.getRange(nextRow, TX.DATUM).setValue(datumWert).setNumberFormat("DD.MM.YYYY");
  wsTx.getRange(nextRow, TX.TYP).setValue("Gutschrift");
  wsTx.getRange(nextRow, TX.DESC).setValue(beschreibung || "Gutschrift");
  wsTx.getRange(nextRow, TX.BETRAG).setValue(betrag).setNumberFormat('#,##0.00 "€"');
  wsTx.getRange(nextRow, TX.GUT_VON).setValue(gutVon);
  wsTx.getRange(nextRow, TX.GUT_AN).setValue(gutAn);

  gutschriftMaskeLeeren_(wsGut);
  schuldenBerechnen_(ss);

  const betragText = Number(betrag).toLocaleString("de-DE", {
    minimumFractionDigits: 2, maximumFractionDigits: 2
  }) + " €";

  ui.alert("✅ Gespeichert!",
    `Gutschrift in Zeile ${nextRow} eingetragen.\n\n` +
    `Von: ${gutVon} → An: ${gutAn}\nBetrag: ${betragText}`,
    ui.ButtonSet.OK);
  wsGut.getRange(G.DATUM).activate();
}


// ════════════════════════════════════════════════════════════════════════
// SCHULDEN-ALGORITHMUS
// ════════════════════════════════════════════════════════════════════════
function schuldenBerechnen() {
  schuldenBerechnen_(SpreadsheetApp.getActiveSpreadsheet());
  SpreadsheetApp.getUi().alert(
    "✅ Neu berechnet",
    "Das Schulden-Sheet wurde aktualisiert.",
    SpreadsheetApp.getUi().ButtonSet.OK
  );
}

function schuldenBerechnen_(ss) {
  const wsBal = ss.getSheetByName("Bilanz");
  const wsSch = ss.getSheetByName(S.SHEET);
  if (!wsBal || !wsSch) return;

  const data = wsBal.getRange(4, 1, 12, 6).getValues();
  const personen = [];
  for (const row of data) {
    const name = row[0];
    const saldo = Number(row[5]) || 0;
    if (name && typeof name === "string" && name.trim() !== "" && name !== "—") {
      personen.push({ name: name.trim(), saldo: Math.round(saldo * 100) / 100 });
    }
  }

  const schuldner = personen
    .filter(p => p.saldo < -0.005)
    .map(p => ({ name: p.name, offen: -p.saldo }))
    .sort((a, b) => b.offen - a.offen);
  const glaeubiger = personen
    .filter(p => p.saldo > 0.005)
    .map(p => ({ name: p.name, offen: p.saldo }))
    .sort((a, b) => b.offen - a.offen);

  const transaktionen = [];
  let si = 0, gi = 0;
  while (si < schuldner.length && gi < glaeubiger.length) {
    const s = schuldner[si];
    const g = glaeubiger[gi];
    const betrag = Math.round(Math.min(s.offen, g.offen) * 100) / 100;
    if (betrag > 0) transaktionen.push({ von: s.name, an: g.name, betrag });
    s.offen -= betrag;
    g.offen -= betrag;
    if (s.offen < 0.005) si++;
    if (g.offen < 0.005) gi++;
  }

  wsSch.getRange(S.DATA_START, 1, S.DATA_MAX - S.DATA_START + 1, 5).clearContent();
  for (let i = 0; i < transaktionen.length && i < 20; i++) {
    const r = S.DATA_START + i;
    const t = transaktionen[i];
    wsSch.getRange(r, 1).setValue(t.von);
    wsSch.getRange(r, 2).setValue(t.an);
    wsSch.getRange(r, 3).setValue(t.betrag).setNumberFormat('#,##0.00 "€"');
    wsSch.getRange(r, 4).setValue("☐");
  }

  const now = new Date();
  const dateStr = Utilities.formatDate(now, Session.getScriptTimeZone(), "dd.MM.yyyy HH:mm");
  wsSch.getRange(S.TIMESTAMP, 1).setValue(dateStr + "  (" + transaktionen.length + " Überweisungen)");
}


// ════════════════════════════════════════════════════════════════════════
// Hilfsfunktionen
// ════════════════════════════════════════════════════════════════════════
function freieZeileFinden_(wsTx) {
  // Nächste freie Zeile in Transaktionen — leer in DESC UND GUT_VON
  const descCol = wsTx.getRange(TX.START, TX.DESC,    TX.MAX - TX.START + 1, 1).getValues();
  const gutCol  = wsTx.getRange(TX.START, TX.GUT_VON, TX.MAX - TX.START + 1, 1).getValues();
  let nextRow = TX.START;
  for (let i = 0; i < descCol.length; i++) {
    if (descCol[i][0] !== "" || gutCol[i][0] !== "") {
      nextRow = TX.START + i + 1;
    }
  }
  return nextRow > TX.MAX ? -1 : nextRow;
}

function alertMissing_(ui, ws, cell, msg) {
  ui.alert("⚠️ Eingabe fehlt", msg, ui.ButtonSet.OK);
  ws.getRange(cell).activate();
}

function eingabeLeeren() {
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  const wsEin = ss.getSheetByName(E.SHEET);
  const ui = SpreadsheetApp.getUi();
  const antwort = ui.alert("Eingabe leeren?", "Alle Felder zurücksetzen?", ui.ButtonSet.YES_NO);
  if (antwort === ui.Button.YES) {
    eingabeMaskeLeeren_(wsEin);
    wsEin.getRange(E.DATUM).activate();
  }
}

function gutschriftLeeren() {
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  const wsGut = ss.getSheetByName(G.SHEET);
  const ui = SpreadsheetApp.getUi();
  const antwort = ui.alert("Gutschrift leeren?", "Alle Felder zurücksetzen?", ui.ButtonSet.YES_NO);
  if (antwort === ui.Button.YES) {
    gutschriftMaskeLeeren_(wsGut);
    wsGut.getRange(G.DATUM).activate();
  }
}

function eingabeMaskeLeeren_(wsEin) {
  wsEin.getRange(E.DATUM).setValue(new Date()).setNumberFormat("DD.MM.YYYY");  // Autofill heute
  wsEin.getRange(E.BESCHREIBUNG).setValue("");
  wsEin.getRange(E.KATEGORIE).setValue("");
  wsEin.getRange(E.BEZAHLT_VON).setValue("");
  wsEin.getRange(E.BETRAG).setValue("");
  wsEin.getRange(E.ALKOHOL).setValue("");
  wsEin.getRange(E.AUFTEILUNG).setValue("Gleichmäßig");
  for (let i = 0; i < E.DABEI_COUNT; i++) {
    wsEin.getRange(E.DABEI_START_ROW + i, E.DABEI_COL).setValue("");
  }
  toggleIndividuellBlock_(wsEin);
}

function gutschriftMaskeLeeren_(wsGut) {
  wsGut.getRange(G.DATUM).setValue(new Date()).setNumberFormat("DD.MM.YYYY");  // Autofill heute
  wsGut.getRange(G.BESCHREIBUNG).setValue("");
  wsGut.getRange(G.BETRAG).setValue("");
  wsGut.getRange(G.GUT_VON).setValue("");
  wsGut.getRange(G.GUT_AN).setValue("");
}


// ════════════════════════════════════════════════════════════════════════
// Button-Funktionen (für gezeichnete Schaltflächen)
// ════════════════════════════════════════════════════════════════════════
function buttonSpeichern()           { transaktionSpeichern(); }
function buttonGutschriftSpeichern() { gutschriftSpeichern(); }
function buttonLeeren()              { eingabeLeeren(); }
function buttonGutschriftLeeren()    { gutschriftLeeren(); }
function buttonSchuldenBerechnen()   { schuldenBerechnen(); }


// ════════════════════════════════════════════════════════════════════════
// Menü + Initial-Zustand beim Öffnen
// ════════════════════════════════════════════════════════════════════════
function onOpen() {
  SpreadsheetApp.getUi()
    .createMenu("⚓ Bordkasse")
    .addItem("💾 Ausgabe speichern",      "transaktionSpeichern")
    .addItem("💾 Gutschrift speichern",   "gutschriftSpeichern")
    .addSeparator()
    .addItem("🔄 Schulden neu berechnen", "schuldenBerechnen")
    .addSeparator()
    .addItem("↺ Eingabe leeren",          "eingabeLeeren")
    .addItem("↺ Gutschrift leeren",       "gutschriftLeeren")
    .addToUi();

  // Initial-Zustand: Individuell-Block am aktuellen Status ausrichten,
  // Datum-Felder auf heute vorbelegen wenn leer.
  const ss = SpreadsheetApp.getActiveSpreadsheet();
  const wsEin = ss.getSheetByName(E.SHEET);
  const wsGut = ss.getSheetByName(G.SHEET);
  if (wsEin) {
    toggleIndividuellBlock_(wsEin);
    if (!wsEin.getRange(E.DATUM).getValue()) {
      wsEin.getRange(E.DATUM).setValue(new Date()).setNumberFormat("DD.MM.YYYY");
    }
  }
  if (wsGut && !wsGut.getRange(G.DATUM).getValue()) {
    wsGut.getRange(G.DATUM).setValue(new Date()).setNumberFormat("DD.MM.YYYY");
  }
}
"""


def write_apps_script():
    DST_JS.write_text(APPS_SCRIPT_V9, encoding="utf-8")
    print(f"✅ Apps Script geschrieben: {DST_JS.relative_to(ROOT)}")


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main():
    if not SRC_XLSX.exists():
        raise SystemExit(f"❌ Quelldatei nicht gefunden: {SRC_XLSX}")

    print(f"→ Lese {SRC_XLSX.relative_to(ROOT)}")
    wb = openpyxl.load_workbook(SRC_XLSX)

    print("→ Baue Eingabe-Tab um (entferne Art-Block + Gutschrift-Block)")
    restructure_eingabe(wb)

    print("→ Erstelle neuen Gutschrift-Tab")
    build_gutschrift(wb)

    print(f"→ Schreibe {DST_XLSX.relative_to(ROOT)}")
    wb.save(DST_XLSX)

    write_apps_script()

    print("\n✅ Migration v8 → v9 abgeschlossen.")
    print("\nNächste Schritte:")
    print(f"  1. Datei {DST_XLSX.name} in Google Sheets öffnen / hochladen")
    print(f"  2. Apps Script-Editor öffnen (Erweiterungen → Apps Script)")
    print(f"  3. Inhalt von {DST_JS.name} dort einfügen, alten Code ersetzen")
    print(f"  4. Speichern & Tabelle neu laden")
    print(f"  5. Test: Ausgabe + Gutschrift erfassen, Aufteilung 'Individuell' togglen")


if __name__ == "__main__":
    main()
